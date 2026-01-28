from datetime import datetime
from io import BytesIO
from html import escape

from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse, HttpResponseBadRequest
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

from database.models import ClassRoom, AttendanceSummary
from ..utils import class_sort_key
from .auth import deny_substitute_access, is_deputy


def _build_daily_export_rows(day):
    classes = sorted(ClassRoom.objects.all(), key=class_sort_key)
    summaries = AttendanceSummary.objects.filter(date=day).select_related('class_room').prefetch_related(
        'absent_students__student')
    summary_by_class = {s.class_room_id: s for s in summaries}

    rows = []
    for class_room in classes:
        summary = summary_by_class.get(class_room.id)
        if not summary:
            rows.append({
                'class_name': class_room.name, 'present_count_reported': '-', 'unexcused_count': '-',
                'unexcused_students': 'Нет данных', 'orvi_count': '-', 'orvi_students': 'Нет данных',
                'other_disease_count': '-', 'other_disease_students': 'Нет данных', 'family_count': '-',
                'family_students': 'Нет данных', 'all_absent_students': 'Нет данных', 'has_data': False,
            })
            continue

        absents = list(summary.absent_students.all())
        has_absents = bool(absents)
        by_reason = {'unexcused': [], 'orvi': [], 'other_disease': [], 'family': []}
        for absent in absents:
            if absent.reason in by_reason:
                by_reason[absent.reason].append(absent.student.full_name)

        all_absent = [absent.student.full_name for absent in absents]

        def format_names(names):
            return ', '.join(names) if has_absents and names else ('Нет данных' if not has_absents else '')

        rows.append({
            'class_name': class_room.name,
            'present_count_reported': summary.present_count_reported,
            'unexcused_count': summary.unexcused_absent_count,
            'unexcused_students': format_names(by_reason['unexcused']),
            'orvi_count': summary.orvi_count,
            'orvi_students': format_names(by_reason['orvi']),
            'other_disease_count': summary.other_disease_count,
            'other_disease_students': format_names(by_reason['other_disease']),
            'family_count': summary.family_reason_count,
            'family_students': format_names(by_reason['family']),
            'all_absent_students': format_names(all_absent),
            'has_data': True,
        })
    return rows


def _export_daily_excel(day, rows):
    headers = ['Класс', 'Пришло', 'Неув.', 'Ученики (неув.)', 'ОРВИ', 'Ученики (ОРВИ)',
               'Другие', 'Ученики (другие)', 'Семейные', 'Ученики (сем.)', 'Все отсутствующие']
    data_keys = ['class_name', 'present_count_reported', 'unexcused_count', 'unexcused_students', 'orvi_count',
                 'orvi_students', 'other_disease_count', 'other_disease_students', 'family_count',
                 'family_students', 'all_absent_students']
    wrap_cols = {4, 6, 8, 10, 11}

    wb = Workbook()
    ws = wb.active
    ws.title = day.strftime('%d.%m.%Y')
    ws.append(headers)

    header_font = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    ok_fill = PatternFill(fill_type='solid', start_color='C6EFCE', end_color='C6EFCE')
    miss_fill = PatternFill(fill_type='solid', start_color='FFC7CE', end_color='FFC7CE')

    for row_idx, row in enumerate(rows, start=2):
        ws.append([row[key] for key in data_keys])
        fill = ok_fill if row['has_data'] else miss_fill
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=(col_idx in wrap_cols), vertical='top')

    widths = {1: 12, 2: 10, 3: 10, 4: 38, 5: 8, 6: 38, 7: 10, 8: 38, 9: 10, 10: 38, 11: 42}
    for col_idx, width in widths.items(): ws.column_dimensions[chr(64 + col_idx)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.getvalue(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="daily_statistics_{day.strftime("%Y-%m-%d")}.xlsx"'
    return response


def _export_daily_word(day, rows):
    headers = ['Класс', 'Пришло', 'Неув.', 'Ученики (неув.)', 'ОРВИ', 'Ученики (ОРВИ)',
               'Другие', 'Ученики (другие)', 'Семейные', 'Ученики (сем.)', 'Все отсутствующие']
    data_keys = ['class_name', 'present_count_reported', 'unexcused_count', 'unexcused_students', 'orvi_count',
                 'orvi_students', 'other_disease_count', 'other_disease_students', 'family_count',
                 'family_students', 'all_absent_students']

    lines = [
        '<!DOCTYPE html><html><head><meta charset="utf-8"><style>',
        '@page WordSection1{size:29.7cm 21.0cm;mso-page-orientation:landscape;margin:1cm;}',
        'div.WordSection1{page:WordSection1;} body{font-family:"Times New Roman",serif;font-size:14pt;}',
        'table{border-collapse:collapse;width:100%;table-layout:fixed;}',
        'th,td{border:1px solid #444;padding:4px;vertical-align:top;font-size:14pt;word-wrap:break-word;}',
        'th{background:#f1f1f1;} .row-ok{background:#e6f4ea;} .row-miss{background:#fde7e9;}',
        '</style></head><body><div class="WordSection1">',
        f'<h2>Дневная статистика за {escape(day.strftime("%d.%m.%Y"))}</h2>',
        '<table><thead><tr>' + ''.join(f'<th>{escape(h)}</th>' for h in headers) + '</tr></thead><tbody>'
    ]
    for row in rows:
        row_class = 'row-ok' if row['has_data'] else 'row-miss'
        lines.append(
            f'<tr class="{row_class}">' + ''.join(f'<td>{escape(str(row[key]))}</td>' for key in data_keys) + '</tr>')
    lines.extend(['</tbody></table></div></body></html>'])

    response = HttpResponse('\n'.join(lines), content_type='application/msword; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="daily_statistics_{day.strftime("%Y-%m-%d")}.doc"'
    return response


@login_required
@deny_substitute_access
@user_passes_test(is_deputy)
def export_daily_statistics(request):
    date_raw = (request.GET.get('date') or '').strip()
    fmt = (request.GET.get('format') or '').strip().lower()

    try:
        day = datetime.strptime(date_raw, '%Y-%m-%d').date() if date_raw else None
    except ValueError:
        day = None

    if not day: return HttpResponseBadRequest('Некорректная дата.')
    if fmt not in ('excel', 'word'): return HttpResponseBadRequest('Некорректный формат.')

    rows = _build_daily_export_rows(day)
    return _export_daily_excel(day, rows) if fmt == 'excel' else _export_daily_word(day, rows)
