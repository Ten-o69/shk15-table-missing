import os
import sys
import re

import django
from openpyxl import load_workbook

# === Django setup ===

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# чтобы Python точно видел твой проект
sys.path.insert(0, BASE_DIR)

# ВАЖНО: если у тебя проект называется иначе,
# поменяй 'shk15_table_missing.settings' на свой модуль настроек.
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "school_attendance.settings")

django.setup()

from database.models import ClassRoom, Student  # твои модели


EXCEL_PATH = os.path.join(BASE_DIR, "classes_students_final_val.xlsx")
SHEET_NAME = ""  # если оставить пустым, будет взят первый лист


def split_students(raw_text: str) -> list[str]:
    """Split cell text into individual students."""
    if not isinstance(raw_text, str):
        return []
    result: list[str] = []
    for line in raw_text.splitlines():
        line = line.strip()
        if not line:
            continue
        parts = re.split(r"\s{2,}", line)
        for part in parts:
            part = part.strip()
            if part:
                result.append(part)
    return result


def import_classes_and_students():
    print("Import started")
    if not os.path.exists(EXCEL_PATH):
        print("Excel file not found:", EXCEL_PATH)
        return

    wb = load_workbook(EXCEL_PATH)
    print("Sheets:", wb.sheetnames)

    if SHEET_NAME and SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
        print("Using sheet:", SHEET_NAME)
    elif SHEET_NAME:
        print("Sheet", repr(SHEET_NAME), "not found, aborting")
        return
    else:
        first = wb.sheetnames[0]
        ws = wb[first]
        print("Using first sheet:", first)

    # First row is header
    header = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    print("Header row:", header)

    created_classes = 0
    created_students = 0

    from django.db import transaction

    with transaction.atomic():
        for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            class_name, students_raw, count_raw, teacher_name = row

            if class_name is None and students_raw is None:
                continue

            class_name = (class_name or "").strip()
            if not class_name:
                print("Row", row_index, ": empty class name, skipping")
                continue

            students_count_expected = None
            if isinstance(count_raw, (int, float)):
                students_count_expected = int(count_raw)

            students_list = split_students(students_raw or "")
            real_count = len(students_list)

            if students_count_expected is not None and students_count_expected != real_count:
                print(
                    "Row", row_index, ", class", repr(class_name),
                    ": expected", students_count_expected, "parsed", real_count
                )

            student_count_value = students_count_expected or real_count or 0

            class_defaults = {"student_count": student_count_value}
            class_room, created = ClassRoom.objects.get_or_create(
                name=class_name,
                defaults=class_defaults,
            )
            if created:
                created_classes += 1
            else:
                if class_room.student_count != student_count_value:
                    class_room.student_count = student_count_value
                    class_room.save()

            for full_name in students_list:
                full_name = full_name.strip()
                if not full_name:
                    continue

                obj, s_created = Student.objects.get_or_create(
                    full_name=full_name,
                    class_room=class_room,
                    defaults={"is_active": True},
                )
                if s_created:
                    created_students += 1

    print("Import finished")
    print("Classes created:", created_classes)
    print("Students created:", created_students)


if __name__ == "__main__":
    import_classes_and_students()
